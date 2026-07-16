#!/usr/bin/env python3
"""
Prepare Learn & Teach resources for php artisan resources:import.

Reads the multi-sheet metadata workbook, creates the folder layout expected by
ResourcesImport (metadata.csv + images/ + links/{group_name}/), and optionally
copies PDFs/thumbnails from a locally synced SharePoint folder.

Usage:
  python3 scripts/prepare-learn-teach-upload.py
  python3 scripts/prepare-learn-teach-upload.py --source-dir "/path/to/New resources 24 March 2026 for LOCALISATION"
  python3 scripts/prepare-learn-teach-upload.py --xlsx "docs/internal/New Resources Learn&Teach Metadata - 23 June 2026.xlsx"

SharePoint sync tip (macOS/Windows):
  In SharePoint, open the folder
  "Ready to upload content / New resources 24 March 2026 for LOCALISATION"
  and choose Sync. Then pass that local path as --source-dir.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
import sys
import urllib.parse
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Install openpyxl: python3 -m pip install openpyxl", file=sys.stderr)
    sys.exit(1)

REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = REPO_ROOT / "docs/internal/New Resources Learn&Teach Metadata - 23 June 2026.xlsx"
DEFAULT_OUT = REPO_ROOT / "docs/internal/learn-teach-upload-2026"

# SharePoint folder names (01–10 numbered; 11–12 use Nature's Codebreakers names).
SHEET_SOURCE_FOLDERS: dict[str, list[str]] = {
    "Binary Number Challenge": ["01 Binary Number Challenge"],
    "Code a Friend": ["02 Code a Friend"],
    "Code words match": ["03 Code Words Match", "03 Code words match"],
    "Coding a binary picture": ["04 Coding a Binary Picture", "04 Coding a binary picture"],
    "Conditional statements": ["05 Conditional statements", "05 Conditional statements - Simon says"],
    "Debugging with a maze": ["06 Debugging with a Maze", "06 Debugging with a maze"],
    "Functions and procedures": ["07 Functions and procedures"],
    "Kitchen algorithms": ["08 Kitchen Algorithm", "08 Kitchen algorithms"],
    "Loop command": ["09 Loop Command", "09 Loop command"],
    "Sorting networks": ["10 Sorting Networks", "10 Sorting networks"],
    "Natures codebreakers1": [
        "11 Nature's Codebreakers part 1",
        "11 Natures codebreakers1",
        "Nature's Codebreakers part 1",
    ],
    "Natures codebreakers2": [
        "12 Nature's Codebreakers part 2",
        "12 Natures codebreakers2",
        "Nature's Codebreakers part 2",
    ],
}

THUMBNAIL_SOURCE_FOLDERS = [
    "Third-party thumbnail",
    "Third party thumbnail",
    "thumbnails",
    "images",
]

CSV_COLUMNS = [
    "name_of_the_resource",
    "link",
    "description",
    "filters_type",
    "filters_target_audience",
    "filters_level_of_difficulty",
    "filters_programming_language",
    "filters_subject",
    "filters_topics",
    "filters_language",
    "category",
    "group_name",
    "image",
    "source_url",
]


def slug_sheet_name(name: str) -> str:
    return name.strip()


def link_to_filename(link: str) -> tuple[str, str]:
    """Return (import_link, source_url). import_link is a local PDF basename or external URL."""
    link = (link or "").strip()
    if not link:
        return "", ""
    if link.startswith("http://") or link.startswith("https://"):
        parsed = urllib.parse.urlparse(link)
        basename = urllib.parse.unquote(Path(parsed.path).name)
        if basename.lower().endswith(".pdf"):
            return basename, link
        return link, link
    return link, ""


def normalize_cell(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def first_image_filename(image: str) -> str:
    """Some rows list multiple SharePoint image URLs separated by ';'."""
    image = (image or "").strip()
    if not image:
        return ""
    if image.startswith("http://") or image.startswith("https://"):
        if ";" in image:
            image = image.split(";")[0].strip()
        parsed = urllib.parse.urlparse(image)
        return urllib.parse.unquote(Path(parsed.path).name)
    return image


def parse_workbook(xlsx_path: Path) -> list[dict]:
    # Must not use read_only/values_only — third-party Link cells store the real URL as a hyperlink.
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    rows_out: list[dict] = []

    for sheet_name in wb.sheetnames:
        group = slug_sheet_name(sheet_name)
        is_third_party = group.lower() == "third party resources"
        ws = wb[sheet_name]

        # Resource sheets: row 2 is an English group blurb (no language); data from row 3.
        # Third-party sheet: data starts at row 2.
        start_row = 2 if is_third_party else 3
        for row_idx, row in enumerate(ws.iter_rows(min_row=start_row), start=start_row):
            values = [c.value for c in row]
            if not any(c is not None and str(c).strip() for c in values):
                continue

            name = normalize_cell(values[2] if len(values) > 2 else "")
            link_raw = normalize_cell(values[3] if len(values) > 3 else "")
            if not name:
                continue

            link_cell = row[3] if len(row) > 3 else None
            if link_cell is not None and link_cell.hyperlink and link_cell.hyperlink.target:
                target = str(link_cell.hyperlink.target).strip()
                if target.startswith("http://") or target.startswith("https://"):
                    link_raw = target

            if is_third_party:
                lang = normalize_cell(values[11] if len(values) > 11 else "")
            else:
                lang = normalize_cell(values[1] if len(values) > 1 else "")
                if not lang or lang.lower() == "language":
                    continue

            import_link, source_url = link_to_filename(link_raw)
            image_raw = normalize_cell(values[12] if len(values) > 12 else "")
            image = first_image_filename(image_raw)
            if image_raw.startswith("http") and not image:
                image = image_raw.split(";")[0].strip()

            rows_out.append(
                {
                    "name_of_the_resource": name,
                    "link": import_link,
                    "description": normalize_cell(values[4] if len(values) > 4 else ""),
                    "filters_type": normalize_cell(values[5] if len(values) > 5 else ""),
                    "filters_target_audience": normalize_cell(values[6] if len(values) > 6 else ""),
                    "filters_level_of_difficulty": normalize_cell(values[7] if len(values) > 7 else ""),
                    "filters_programming_language": normalize_cell(values[8] if len(values) > 8 else ""),
                    "filters_subject": normalize_cell(values[9] if len(values) > 9 else ""),
                    "filters_topics": normalize_cell(values[10] if len(values) > 10 else ""),
                    "filters_language": lang if not is_third_party else normalize_cell(values[11] if len(values) > 11 else ""),
                    "category": "",
                    "group_name": group if not is_third_party else "Third party resources",
                    "image": image,
                    "source_url": source_url or (link_raw if link_raw.startswith("http") else ""),
                    "_link_raw": link_raw,
                    "_image_raw": image_raw,
                    "_needs_pdf": bool(import_link.lower().endswith(".pdf")),
                    "_is_third_party": is_third_party,
                }
            )

    return rows_out


def build_file_index(source_dir: Path) -> dict[str, Path]:
    index: dict[str, Path] = {}
    for path in source_dir.rglob("*"):
        if not path.is_file():
            continue
        key = path.name.casefold()
        if key not in index:
            index[key] = path
    return index


def find_in_source(source_dir: Path, filename: str, preferred_subdirs: list[str] | None = None) -> Path | None:
    if not filename or not source_dir.is_dir():
        return None

    if preferred_subdirs:
        for sub in preferred_subdirs:
            candidate = source_dir / sub / filename
            if candidate.is_file():
                return candidate
            for match in (source_dir / sub).rglob(filename):
                if match.is_file():
                    return match

    direct = source_dir / filename
    if direct.is_file():
        return direct

    folded = filename.casefold()
    for path in source_dir.rglob("*"):
        if path.is_file() and path.name.casefold() == folded:
            return path
    return None


def copy_assets(rows: list[dict], source_dir: Path, out_dir: Path) -> dict:
    images_dir = out_dir / "images"
    links_dir = out_dir / "links"
    images_dir.mkdir(parents=True, exist_ok=True)
    links_dir.mkdir(parents=True, exist_ok=True)

    stats = {
        "pdfs_found": 0,
        "pdfs_missing": 0,
        "images_found": 0,
        "images_missing": 0,
        "missing_pdfs": [],
        "missing_images": [],
    }

    seen_images: set[str] = set()

    for row in rows:
        group = row["group_name"]
        group_links = links_dir / group
        group_links.mkdir(parents=True, exist_ok=True)

        if row["_needs_pdf"]:
            pdf_name = row["link"]
            preferred = SHEET_SOURCE_FOLDERS.get(group.strip(), [])
            src = find_in_source(source_dir, pdf_name, preferred)
            dest = group_links / pdf_name
            if src and src.is_file():
                if not dest.exists() or dest.stat().st_size != src.stat().st_size:
                    shutil.copy2(src, dest)
                stats["pdfs_found"] += 1
            else:
                stats["pdfs_missing"] += 1
                stats["missing_pdfs"].append({"group": group, "file": pdf_name, "source_url": row["source_url"]})

        image = row["image"]
        if not image or image.startswith("http") or image in seen_images:
            continue
        seen_images.add(image)

        preferred_thumb = THUMBNAIL_SOURCE_FOLDERS + SHEET_SOURCE_FOLDERS.get(group.strip(), [])
        src = find_in_source(source_dir, image, preferred_thumb)
        dest = images_dir / image
        if src and src.is_file():
            if not dest.exists() or dest.stat().st_size != src.stat().st_size:
                shutil.copy2(src, dest)
            stats["images_found"] += 1
        else:
            stats["images_missing"] += 1
            stats["missing_images"].append({"group": group, "file": image, "raw": row["_image_raw"]})

    return stats


def write_csv(rows: list[dict], out_path: Path) -> None:
    with out_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k, "") for k in CSV_COLUMNS})


def write_manifest(rows: list[dict], stats: dict | None, out_path: Path) -> None:
    payload = {
        "total_rows": len(rows),
        "pdf_rows": sum(1 for r in rows if r["_needs_pdf"]),
        "third_party_rows": sum(1 for r in rows if r["_is_third_party"]),
        "third_party_missing_urls": [
            {"name": r["name_of_the_resource"], "link": r["_link_raw"]}
            for r in rows
            if r["_is_third_party"] and not r["link"].startswith("http")
        ],
        "copy_stats": stats,
        "rows": [
            {
                "group_name": r["group_name"],
                "name_of_the_resource": r["name_of_the_resource"],
                "filters_language": r["filters_language"],
                "link": r["link"],
                "image": r["image"],
                "needs_pdf": r["_needs_pdf"],
                "source_url": r["source_url"],
            }
            for r in rows
        ],
    }
    out_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description="Prepare Learn & Teach upload folder for resources:import")
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX, help="Path to metadata workbook")
    parser.add_argument("--out", type=Path, default=DEFAULT_OUT, help="Output directory")
    parser.add_argument(
        "--languages",
        type=str,
        default="",
        help="Comma-separated list of language names from the sheet (e.g. Albanian,Bulgarian). "
        "When set, only resource rows for these languages are included. Third-party rows are included unless --exclude-third-party is set.",
    )
    parser.add_argument(
        "--exclude-third-party",
        action="store_true",
        help="Exclude the 'Third party resources' sheet rows from metadata.csv (useful when you only want localized PDFs).",
    )
    parser.add_argument(
        "--source-dir",
        type=Path,
        default=None,
        help="Locally synced SharePoint folder containing PDFs and thumbnails",
    )
    args = parser.parse_args()

    if not args.xlsx.is_file():
        print(f"Workbook not found: {args.xlsx}", file=sys.stderr)
        return 1

    rows = parse_workbook(args.xlsx)

    wanted_languages: set[str] = set()
    if isinstance(args.languages, str) and args.languages.strip():
        wanted_languages = {normalize_cell(x) for x in args.languages.split(",") if normalize_cell(x)}

    if args.exclude_third_party or wanted_languages:
        filtered: list[dict] = []
        for row in rows:
            if args.exclude_third_party and row["_is_third_party"]:
                continue
            if wanted_languages and not row["_is_third_party"]:
                if normalize_cell(str(row.get("filters_language", ""))) not in wanted_languages:
                    continue
            filtered.append(row)
        rows = filtered

    print(f"Using {len(rows)} metadata rows"
          + (f" (languages: {', '.join(sorted(wanted_languages))})" if wanted_languages else ""))
    args.out.mkdir(parents=True, exist_ok=True)

    stats = None
    if args.source_dir:
        if not args.source_dir.is_dir():
            print(f"Source directory not found: {args.source_dir}", file=sys.stderr)
            return 1
        stats = copy_assets(rows, args.source_dir, args.out)
    else:
        (args.out / "images").mkdir(parents=True, exist_ok=True)
        (args.out / "links").mkdir(parents=True, exist_ok=True)
        for row in rows:
            if row["_needs_pdf"]:
                (args.out / "links" / row["group_name"]).mkdir(parents=True, exist_ok=True)

    write_csv(rows, args.out / "metadata.csv")
    write_manifest(rows, stats, args.out / "manifest.json")

    print(f"Wrote {len(rows)} rows to {args.out / 'metadata.csv'}")
    print(f"Manifest: {args.out / 'manifest.json'}")
    print(f"Folder layout: {args.out}/{{metadata.csv, images/, links/<group>/}}")

    if stats:
        print(
            f"PDFs: {stats['pdfs_found']} copied, {stats['pdfs_missing']} missing | "
            f"Images: {stats['images_found']} copied, {stats['images_missing']} missing"
        )
        if stats["pdfs_missing"] or stats["images_missing"]:
            print("See manifest.json copy_stats for missing file details.")
            return 2

    missing_urls = [r for r in rows if r["_is_third_party"] and not r["link"].startswith("http")]
    if missing_urls:
        print(f"Warning: {len(missing_urls)} third-party rows still need external URLs in metadata.csv")

    if not args.source_dir:
        print("\nNext: sync SharePoint locally, then re-run with --source-dir to copy PDFs and images.")

    return 0


if __name__ == "__main__":
    sys.exit(main())
