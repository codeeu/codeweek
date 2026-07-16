#!/usr/bin/env python3
"""
Download Learn & Teach PDFs/thumbnails from SharePoint using Azure CLI auth.

Prereq:
  az login --tenant dc75c1d7-1a32-4e97-af90-1dc01911ea6c --allow-no-subscriptions

Usage:
  python3 scripts/download-learn-teach-from-sharepoint.py
  python3 scripts/download-learn-teach-from-sharepoint.py --languages Albanian
"""

from __future__ import annotations

import argparse
import json
import subprocess
import sys
import time
import urllib.parse
import urllib.request
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Install openpyxl first", file=sys.stderr)
    sys.exit(1)

REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = REPO_ROOT / "docs/internal/New Resources Learn&Teach Metadata - 23 June 2026.xlsx"
DEFAULT_OUT = REPO_ROOT / "docs/internal/learn-teach-upload-2026"
SITE_ID = "jaye.sharepoint.com,9352ca34-6dc1-4d96-b96e-69e1bf73de15,04d001a9-5fb2-4d9b-b0a6-6079b5a75b5e"
GRAPH = "https://graph.microsoft.com/v1.0"


def get_token() -> str:
    result = subprocess.run(
        [
            "az",
            "account",
            "get-access-token",
            "--resource",
            "https://graph.microsoft.com",
            "--query",
            "accessToken",
            "-o",
            "tsv",
        ],
        check=True,
        capture_output=True,
        text=True,
    )
    token = result.stdout.strip()
    if not token:
        raise RuntimeError("Empty Graph token. Run: az login --tenant dc75c1d7-1a32-4e97-af90-1dc01911ea6c --allow-no-subscriptions")
    return token


def graph_get(token: str, url: str) -> dict:
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {token}"})
    with urllib.request.urlopen(req, timeout=90) as resp:
        return json.loads(resp.read().decode("utf-8"))


def graph_download(token: str, url: str, dest: Path) -> None:
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {token}"})
    with urllib.request.urlopen(req, timeout=180) as resp:
        dest.parent.mkdir(parents=True, exist_ok=True)
        dest.write_bytes(resp.read())


def search_file(token: str, filename: str) -> dict | None:
    # OData string literal: escape single quotes by doubling them.
    # Also try a fallback query without apostrophes if needed.
    candidates = [filename]
    if "'" in filename or "’" in filename:
        candidates.append(filename.replace("'", "").replace("’", ""))
        candidates.append(filename.replace("Nature's", "Natures").replace("Nature’s", "Natures"))

    for candidate in candidates:
        literal = candidate.replace("'", "''")
        encoded = urllib.parse.quote(literal, safe="")
        url = f"{GRAPH}/sites/{SITE_ID}/drive/root/search(q='{encoded}')"
        try:
            data = graph_get(token, url)
        except Exception:
            continue
        values = data.get("value") or []
        exact = [v for v in values if (v.get("name") or "").casefold() == filename.casefold()]
        if not exact:
            # Fuzzy: same basename ignoring apostrophes
            norm = lambda s: s.casefold().replace("'", "").replace("’", "")
            exact = [v for v in values if norm(v.get("name") or "") == norm(filename)]
        if exact:
            preferred = [
                v
                for v in exact
                if "LOCALISATION" in (v.get("webUrl") or "")
            ]
            return preferred[0] if preferred else exact[0]
    return None


def link_to_filename(link: str) -> str:
    link = (link or "").strip()
    if not link:
        return ""
    if link.startswith("http://") or link.startswith("https://"):
        path = urllib.parse.urlparse(link).path
        name = urllib.parse.unquote(Path(path).name)
        return name if name.lower().endswith(".pdf") else ""
    return link if link.lower().endswith(".pdf") else ""


def first_image_name(image: str) -> str:
    image = (image or "").strip()
    if not image:
        return ""
    if image.startswith("http"):
        if ";" in image:
            image = image.split(";")[0].strip()
        path = urllib.parse.urlparse(image).path
        return urllib.parse.unquote(Path(path).name)
    return image


def parse_needed_assets(xlsx: Path, languages: set[str]) -> tuple[list[tuple[str, str]], set[str]]:
    """Return ([(group, pdf_filename), ...], image_filenames)."""
    wb = openpyxl.load_workbook(xlsx, read_only=True)
    pdfs: list[tuple[str, str]] = []
    images: set[str] = set()

    for sheet_name in wb.sheetnames:
        group = sheet_name.strip()
        is_third = group.lower() == "third party resources"
        ws = wb[sheet_name]
        start = 2 if is_third else 3
        for row in ws.iter_rows(min_row=start, values_only=True):
            if not any(c is not None and str(c).strip() for c in row):
                continue
            name = str(row[2]).strip() if row[2] else ""
            if not name:
                continue
            if is_third:
                lang = str(row[11]).strip() if len(row) > 11 and row[11] else ""
            else:
                lang = str(row[1]).strip() if row[1] else ""
                if not lang or lang.lower() == "language":
                    continue
            if languages and not is_third and lang not in languages:
                continue

            link = str(row[3]).strip() if row[3] else ""
            pdf = link_to_filename(link)
            if pdf and not is_third:
                pdfs.append((group, pdf))

            image = first_image_name(str(row[12]).strip() if len(row) > 12 and row[12] else "")
            if image and not image.startswith("http") and "." in image and image.lower() not in {"view", "view?usp=drive_link"}:
                images.add(image)

    # unique pdfs preserving order
    seen = set()
    unique_pdfs = []
    for item in pdfs:
        if item not in seen:
            seen.add(item)
            unique_pdfs.append(item)
    return unique_pdfs, images


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--out", type=Path, default=DEFAULT_OUT)
    parser.add_argument("--languages", default="", help="Comma-separated, e.g. Albanian,Bulgarian")
    parser.add_argument("--skip-existing", action="store_true", default=True)
    args = parser.parse_args()

    languages = {x.strip() for x in args.languages.split(",") if x.strip()}
    pdfs, images = parse_needed_assets(args.xlsx, languages)
    print(f"Need {len(pdfs)} PDFs and {len(images)} image filenames")

    token = get_token()
    missing_pdfs = []
    missing_images = []
    ok_pdfs = 0
    ok_images = 0

    for i, (group, pdf_name) in enumerate(pdfs, 1):
        dest = args.out / "links" / group / pdf_name
        if args.skip_existing and dest.is_file() and dest.stat().st_size > 1000:
            ok_pdfs += 1
            continue
        try:
            item = search_file(token, pdf_name)
            if not item:
                missing_pdfs.append({"group": group, "file": pdf_name})
                print(f"[{i}/{len(pdfs)}] MISSING {pdf_name}")
                continue
            graph_download(token, f"{GRAPH}/sites/{SITE_ID}/drive/items/{item['id']}/content", dest)
            # sanity check
            if not dest.read_bytes().startswith(b"%PDF"):
                dest.unlink(missing_ok=True)
                missing_pdfs.append({"group": group, "file": pdf_name, "reason": "not a pdf"})
                print(f"[{i}/{len(pdfs)}] BAD PDF {pdf_name}")
                continue
            ok_pdfs += 1
            print(f"[{i}/{len(pdfs)}] OK {group}/{pdf_name}")
            time.sleep(0.15)
        except Exception as e:
            missing_pdfs.append({"group": group, "file": pdf_name, "error": str(e)})
            print(f"[{i}/{len(pdfs)}] ERROR {pdf_name}: {e}")

    for i, image_name in enumerate(sorted(images), 1):
        dest = args.out / "images" / image_name
        if args.skip_existing and dest.is_file() and dest.stat().st_size > 100:
            ok_images += 1
            continue
        try:
            item = search_file(token, image_name)
            if not item:
                missing_images.append({"file": image_name})
                print(f"[img {i}/{len(images)}] MISSING {image_name}")
                continue
            graph_download(token, f"{GRAPH}/sites/{SITE_ID}/drive/items/{item['id']}/content", dest)
            ok_images += 1
            print(f"[img {i}/{len(images)}] OK {image_name}")
            time.sleep(0.15)
        except Exception as e:
            missing_images.append({"file": image_name, "error": str(e)})
            print(f"[img {i}/{len(images)}] ERROR {image_name}: {e}")

    report = {
        "ok_pdfs": ok_pdfs,
        "ok_images": ok_images,
        "missing_pdfs": missing_pdfs,
        "missing_images": missing_images,
    }
    report_path = args.out / "download-report.json"
    report_path.write_text(json.dumps(report, indent=2), encoding="utf-8")
    print(f"\nDone. PDFs {ok_pdfs}/{len(pdfs)}, images {ok_images}/{len(images)}")
    print(f"Report: {report_path}")
    return 0 if not missing_pdfs else 2


if __name__ == "__main__":
    sys.exit(main())
